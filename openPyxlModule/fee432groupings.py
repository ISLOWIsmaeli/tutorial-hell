from openpyxl import load_workbook
filename="groupings.xlsx"
workbook = load_workbook(filename)
mysheets = workbook.sheetnames

class_list_sheet=workbook[mysheets[0]]
projects_without_ismael_sheet=workbook[mysheets[1]]
projects_with_ismael_sheet=workbook[mysheets[2]]

for without_ismael in projects_without_ismael_sheet.iter_rows(min_row=2,
                                                              max_row=120,
                                                              min_col=3,
                                                              max_col=3,
                                                              values_only=True):
    counter=2
    classmate1=without_ismael[0]
    for class_list in class_list_sheet.iter_rows(min_row=3,
                                             max_row=128,
                                             min_col=2,
                                             max_col=2,
                                             values_only=True):
        counter=counter+1
        classmate2=class_list[0]
        if classmate1==classmate2:
            print(classmate1,"has a group at postion",counter,sep=": ")
            class_list_sheet["A{}".format(counter)]=""
            class_list_sheet["B{}".format(counter)]=""
            class_list_sheet["C{}".format(counter)]=""
            class_list_sheet["D{}".format(counter)]=""
            class_list_sheet["E{}".format(counter)]=""

            break
        else:
            pass
workbook.save(filename)
