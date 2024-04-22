import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Replace 'input.xlsx' with the name of your Excel workbook
# Replace 'Sheet1' with the name of the sheet containing the data
excel_file = 'C:\\Users\\Lennox\\Documents\\myProjects\\tutorials\\tutorial-hell\\openPyxlModule\\NUMERICAL METHODS CLASS LIST (1).xlsx'
sheet_name = 'Form Responses 1'

df = pd.read_excel(excel_file, sheet_name=sheet_name)

random.seed()
df = df.sample(frac=1).reset_index(drop=True)
groups = []
group_size = 3

for i in range(0, len(df), group_size):
    groups.append(df.iloc[i:i + group_size])

# Replace 'output.xlsx' with the name of your output Excel workbook (it can be the same as the input workbook)
output_file = 'C:\\Users\\Lennox\\Documents\\myProjects\\tutorials\\tutorial-hell\\openPyxlModule\\NUMERICAL METHODS CLASS LIST (1).xlsx'
output_sheet_name = 'Sheet2'

wb = load_workbook(output_file)

if output_sheet_name in wb.sheetnames:
    ws = wb[output_sheet_name]
else:
    ws = wb.create_sheet(output_sheet_name)

ws.append(["Group", "Registration Number", "Name"])

for i, group in enumerate(groups, start=1):
    for row in dataframe_to_rows(group, index=False, header=False):
        ws.append([f"Group {i}", *row])

wb.save(output_file)
