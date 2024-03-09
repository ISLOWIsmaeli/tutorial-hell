from datetime import datetime
from openpyxl import load_workbook
from classes import Product, Review
from mapping import (PRODUCT_ID,PRODUCT_PARENT,PRODUCT_TITLE, PRODUCT_CATEGORY,
                     REVIEW_DATE,REVIEW_ID,REVIEW_CUSTOMER,REVIEW_STARS,REVIEW_HEADLINE,
                     REVIEW_BODY)

workbook= load_workbook(filename="sample.xlsx",read_only=True)
sheet=workbook.active

products=[]
reviews=[]

for row in sheet.iter_rows(min_row=2,values_only=True):
    product=Product(id=row[PRODUCT_ID],
                    parent=row[PRODUCT_PARENT],
                    title=row[PRODUCT_TITLE],
                    category=row[PRODUCT_CATEGORY])
    products.append(product)
    spread_date=row[REVIEW_DATE]
    parsed_date=datetime.strptime(spread_date,"%Y-%m-%d")
    review=Review(id=row[REVIEW_ID],
                  customer_id=row[REVIEW_CUSTOMER],
                  stars=row[REVIEW_STARS],
                  headline=row[REVIEW_HEADLINE],
                  body=row[REVIEW_BODY],
                  date=parsed_date)
    reviews.append(review)
print(products[0])
print(reviews[0])