from openpyxl import load_workbook
import random

filename="testData.xlsx"
workbook=load_workbook(filename)

sheet1=workbook["Sheet1"]
no_of_sites=int(input("Please enter the number of sites: "))
site_names =[]
missionary_list=[]

for i in range(no_of_sites):
    site_entry=input("Site {} ".format(i+1))
    site_names.append(site_entry)

#to create worksheets for each and every site
    
for missionary in sheet1.iter_rows(min_row=2,
                                   max_row=50,
                                   min_col=1,
                                   max_col=3,
                                   values_only=True):
    #to iterate through all the missionaries and store them in a list
    missionary_list.append(missionary)
    # print(missionary)
print(random.choice(missionary_list))

for missionary in missionary_list:
    pick=random.choice(missionary_list)
    index= missionary_list.index(pick)
    for confirm in missionary_list:
        if pick == confirm:
            del missionary_list[index]
            continue
        elif ValueError:
            continue
print(missionary_list)