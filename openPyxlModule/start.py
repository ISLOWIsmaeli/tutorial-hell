from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# wb = load_workbook('C:\\Users\\Lennox\\documents\\myProjects\\tutorials\\tutorial-hell\\openPyxlModule\\grades.xlsx')
# ws = wb['Grades']
# wb.create_sheet("Test")
# print(wb.sheetnames)
wb = load_workbook('ismael.xlsx')
wb=Workbook()
ws=wb.active
for row in range(1,11):
    for col in range (1,5):
        # char = chr(65 + col) the manual way, use range 1 to 5
        char = get_column_letter(col)
        print(ws[char + str(row)].value)
wb.save('ismael.xlsx')
