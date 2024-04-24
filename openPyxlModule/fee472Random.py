import pandas as pd
import random
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

excel_file = '' # replace with the path to your excel file, remember to use '\\' for windows and '/' for UNIX (mac and linux)
sheet_name = 'Class List'
df = pd.read_excel(excel_file, sheet_name=sheet_name)
random.seed()
df = df.sample(frac=1).reset_index(drop=True)
groups = []
group_size = 3

for i in range(0, len(df), group_size):
    groups.append(df.iloc[i:i + group_size])

output_file = '' # replace with the path to your output excel file, , remember to use '\\' for windows and '/' for UNIX (mac and linux)
output_sheet_name = 'Sorted Class List'
wb = load_workbook(output_file)

if output_sheet_name in wb.sheetnames:
    ws = wb[output_sheet_name]
else:
    ws = wb.create_sheet(output_sheet_name)

ws.append(["Group", "Registration Number", "Name"])
for i, group in enumerate(groups, start=1):
    for row in dataframe_to_rows(group, index=False, header=False):
        ws.append([f"Group {i}", *row])

wb.save(output_file)
